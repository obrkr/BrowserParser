#!/usr/bin/env python3
"""
Chrome-based Browser History Parser
Parses browsing history from Chrome, Edge, Brave, and other Chromium browsers.
Exports to XLSX with tabs: History, Searches, Downloads

Usage: Run the script and it will prompt you for the folder containing the History file.
"""

import os
os.environ["TK_SILENCE_DEPRECATION"] = "1"

import sqlite3
import shutil
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import tkinter as tk
from tkinter import filedialog


def copy_history_db(source_path: str) -> str:
    """Copy the history database to a temp location (Chrome locks the original)."""
    temp_path = os.path.join(os.path.dirname(source_path), "history_copy.db")
    shutil.copy2(source_path, temp_path)
    return temp_path


def parse_history(db_path: str) -> dict:
    """Parse the browser history database and extract all data."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # First, inspect the schema to determine the correct query
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tables = [row[0] for row in cursor.fetchall()]
    
    history = []
    searches = []
    downloads = []
    
    # Parse regular browsing history from urls table
    cursor.execute("PRAGMA table_info(urls)")
    urls_columns = [row[1] for row in cursor.fetchall()]
    
    # Check if visits table exists (newer Chrome schema)
    has_visits = 'visits' in tables
    
    if 'visit_time' in urls_columns:
        cursor.execute("""
            SELECT url, title, visit_time, visit_count
            FROM urls
            ORDER BY visit_time DESC
        """)
    elif has_visits:
        # Check visits table schema
        cursor.execute("PRAGMA table_info(visits)")
        visits_columns = [row[1] for row in cursor.fetchall()]
        
        if 'visit_time' in visits_columns and 'url' in visits_columns:
            cursor.execute("""
                SELECT u.url, u.title, v.visit_time, u.visit_count
                FROM urls u
                LEFT JOIN visits v ON u.id = v.url
                ORDER BY v.visit_time DESC
            """)
        else:
            # Fallback
            cursor.execute("""
                SELECT url, title, visit_count
                FROM urls
                ORDER BY visit_count DESC
            """)
    else:
        cursor.execute("""
            SELECT url, title, visit_count
            FROM urls
            ORDER BY visit_count DESC
        """)
    
    for row in cursor.fetchall():
        if len(row) == 4:
            url, title, visit_time, visit_count = row
        else:
            url, title, visit_count = row
            visit_time = None
        
        timestamp = None
        if visit_time:
            try:
                timestamp = datetime(1601, 1, 1) + timedelta(microseconds=visit_time)
            except:
                pass
        
        url_str = url or ""
        title_str = title or ""
        
        entry = {
            "url": url_str,
            "title": title_str,
            "timestamp": timestamp,
            "visit_count": visit_count,
        }
        
        history.append(entry)
    
    # Parse searches from keyword_search_terms table (usually in WebData file)
    if 'keyword_search_terms' in tables:
        cursor.execute("PRAGMA table_info(keyword_search_terms)")
        search_columns = [row[1] for row in cursor.fetchall()]
        
        if 'timestamp' in search_columns:
            cursor.execute("SELECT term, url_id, timestamp FROM keyword_search_terms ORDER BY timestamp DESC")
            for term, url_id, timestamp in cursor.fetchall():
                ts = None
                if timestamp:
                    try:
                        ts = datetime(1601, 1, 1) + timedelta(microseconds=timestamp)
                    except:
                        pass
                searches.append({
                    "search_term": term or "",
                    "timestamp": ts,
                    "url_id": url_id,
                })
        else:
            # Try to get timestamp from visits table via url_id
            cursor.execute("SELECT term, url_id FROM keyword_search_terms")
            for term, url_id in cursor.fetchall():
                ts = None
                if has_visits and url_id:
                    try:
                        cursor.execute("SELECT visit_time FROM visits WHERE url = ?", (url_id,))
                        row = cursor.fetchone()
                        if row and row[0]:
                            ts = datetime(1601, 1, 1) + timedelta(microseconds=row[0])
                    except:
                        pass
                searches.append({
                    "search_term": term or "",
                    "timestamp": ts,
                    "url_id": url_id,
                })
    
    # Parse downloads from downloads table (usually in History file)
    if 'downloads' in tables:
        cursor.execute("PRAGMA table_info(downloads)")
        download_columns = [row[1] for row in cursor.fetchall()]
        
        # Build query based on available columns
        cols = []
        if 'url' in download_columns:
            cols.append("url")
        if 'start_time' in download_columns:
            cols.append("start_time")
        if 'end_time' in download_columns:
            cols.append("end_time")
        if 'state' in download_columns:
            cols.append("state")
        if 'full_path' in download_columns:
            cols.append("full_path")
        elif 'target_path' in download_columns:
            cols.append("target_path")
        
        if not cols:
            pass  # No columns available, skip downloads
        
        query = f"SELECT {', '.join(cols)} FROM downloads"
        if 'start_time' in cols:
            query += " ORDER BY start_time DESC"
        
        cursor.execute(query)
        rows = cursor.fetchall()
        
        for row in rows:
            row_dict = dict(zip(cols, row))
            
            ts = None
            if 'start_time' in row_dict and row_dict['start_time']:
                try:
                    ts = datetime(1601, 1, 1) + timedelta(microseconds=row_dict['start_time'])
                except:
                    pass
            
            file_path = row_dict.get('full_path') or row_dict.get('target_path') or ""
            
            downloads.append({
                "file_path": file_path,
                "url": row_dict.get('url', ""),
                "timestamp": ts,
                "state": row_dict.get('state', ""),
            })
    
    conn.close()
    return {
        "history": history,
        "searches": searches,
        "downloads": downloads,
        "extensions": []
    }


def parse_extensions(profile_path: str) -> list:
    """Parse browser extensions from the Extensions folder."""
    import json
    from datetime import datetime, timedelta
    
    extensions = []
    
    # Chrome/Edge/Brave store extensions in the Extensions folder as manifest.json files
    ext_folder = os.path.join(profile_path, "Extensions")
    
    if not os.path.exists(ext_folder):
        return extensions
    
    try:
        # Each subfolder is an extension ID
        for ext_id in os.listdir(ext_folder):
            ext_path = os.path.join(ext_folder, ext_id)
            if not os.path.isdir(ext_path):
                continue
            
            # Look for manifest.json in version folders
            for version in os.listdir(ext_path):
                version_path = os.path.join(ext_path, version)
                if not os.path.isdir(version_path):
                    continue
                
                manifest_path = os.path.join(version_path, "manifest.json")
                if os.path.exists(manifest_path):
                    try:
                        with open(manifest_path, 'r', encoding='utf-8') as f:
                            manifest = json.load(f)
                            name = manifest.get('name', '')
                            ext_version = manifest.get('version', '')
                            
                            # Clean up name (remove __MSG_ references)
                            if name.startswith('__MSG_'):
                                name = ext_id  # Use extension ID as fallback
                            
                            # Build Chrome Web Store URL
                            store_url = f"https://chrome.google.com/webstore/detail/{ext_id}"
                            
                            extensions.append({
                                "name": name,
                                "version": ext_version,
                                "url": store_url,
                            })
                            break  # Only take first version
                    except:
                        pass
    except:
        pass
    
    return extensions


def export_to_xlsx(data: dict, output_path: str):
    """Export the parsed data to an Excel file with multiple sheets."""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Create History sheet
    ws_history = wb.create_sheet("History")
    headers_history = ["Timestamp", "Title", "URL", "Visit Count"]
    for col, header in enumerate(headers_history, 1):
        cell = ws_history.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row, entry in enumerate(data["history"], 2):
        ws_history.cell(row=row, column=1, value=entry["timestamp"].strftime("%Y-%m-%d %H:%M:%S UTC") if entry["timestamp"] else "")
        ws_history.cell(row=row, column=2, value=entry["title"])
        ws_history.cell(row=row, column=3, value=entry["url"])
        ws_history.cell(row=row, column=4, value=entry["visit_count"])
    
    # Auto-adjust column widths
    for col in ws_history.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_history.column_dimensions[column].width = min(max_length + 2, 60)
    
    # Create Searches sheet
    ws_searches = wb.create_sheet("Searches")
    headers_searches = ["Timestamp", "Search Term", "URL ID"]
    for col, header in enumerate(headers_searches, 1):
        cell = ws_searches.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row, entry in enumerate(data["searches"], 2):
        ws_searches.cell(row=row, column=1, value=entry["timestamp"].strftime("%Y-%m-%d %H:%M:%S UTC") if entry["timestamp"] else "")
        ws_searches.cell(row=row, column=2, value=entry.get("search_term", ""))
        ws_searches.cell(row=row, column=3, value=entry.get("url_id", ""))
    
    for col in ws_searches.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_searches.column_dimensions[column].width = min(max_length + 2, 60)
    
    # Create Downloads sheet
    ws_downloads = wb.create_sheet("Downloads")
    headers_downloads = ["Timestamp", "File Path", "URL", "State"]
    for col, header in enumerate(headers_downloads, 1):
        cell = ws_downloads.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row, entry in enumerate(data["downloads"], 2):
        ws_downloads.cell(row=row, column=1, value=entry["timestamp"].strftime("%Y-%m-%d %H:%M:%S UTC") if entry["timestamp"] else "")
        ws_downloads.cell(row=row, column=2, value=entry.get("file_path", ""))
        ws_downloads.cell(row=row, column=3, value=entry.get("url", ""))
        ws_downloads.cell(row=row, column=4, value=entry.get("state", ""))
    
    for col in ws_downloads.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_downloads.column_dimensions[column].width = min(max_length + 2, 60)
    
    # Create Extensions sheet
    ws_extensions = wb.create_sheet("Extensions")
    headers_extensions = ["Name", "Version", "Install Date", "Store URL"]
    for col, header in enumerate(headers_extensions, 1):
        cell = ws_extensions.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row, entry in enumerate(data.get("extensions", []), 2):
        ws_extensions.cell(row=row, column=1, value=entry.get("name", ""))
        ws_extensions.cell(row=row, column=2, value=entry.get("version", ""))
        ws_extensions.cell(row=row, column=3, value=entry.get("install_time").strftime("%Y-%m-%d %H:%M:%S UTC") if entry.get("install_time") else "")
        ws_extensions.cell(row=row, column=4, value=entry.get("url", ""))
    
    for col in ws_extensions.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_extensions.column_dimensions[column].width = min(max_length + 2, 60)
    
    # Save workbook
    wb.save(output_path)


def select_history_file():
    """Open a file dialog to let user select the History file or folder."""
    # Hide the root tkinter window
    root = tk.Tk()
    root.withdraw()
    
    print("Please select the folder containing the 'History' file.")
    print()
    print("┌─ macOS ────────────────────────────────────────────────────────────────────────────────")
    print("│ Chrome:    ~/Library/Application Support/Google/Chrome/Default/")
    print("│ Edge:      ~/Library/Application Support/Microsoft Edge/Default/")
    print("│ Brave:     ~/Library/Application Support/BraveSoftware/Brave-Browser/Default/")
    print("│ Opera:     ~/Library/Application Support/com.operasoftware.Opera/Default/")
    print("│ Vivaldi:   ~/Library/Application Support/Vivaldi/Default/")
    print("│ Chromium:  ~/Library/Application Support/Chromium/Default/")
    print("├─ Windows ──────────────────────────────────────────────────────────────────────────────")
    print("│ Chrome:    %LOCALAPPDATA%\\Google\\Chrome\\User Data\\Default\\")
    print("│ Edge:      %LOCALAPPDATA%\\Microsoft\\Edge\\User Data\\Default\\")
    print("│ Brave:     %LOCALAPPDATA%\\BraveSoftware\\Brave-Browser\\User Data\\Default\\")
    print("│ Opera:     %APPDATA%\\Opera Software\\Opera Stable\\")
    print("│ Vivaldi:   %LOCALAPPDATA%\\Vivaldi\\User Data\\Default\\")
    print("│ Chromium:  %LOCALAPPDATA%\\Chromium\\User Data\\Default\\")
    print("├─ Linux ────────────────────────────────────────────────────────────────────────────────")
    print("│ Chrome:    ~/.config/google-chrome/Default/")
    print("│ Edge:      ~/.config/microsoft-edge/Default/")
    print("│ Brave:     ~/.config/BraveSoftware/Brave-Browser/Default/")
    print("│ Chromium:  ~/.config/chromium/Default/")
    print("└────────────────────────────────────────────────────────────────────────────────────────")
    print()
    
    # Ask user to select the folder
    folder_selected = filedialog.askdirectory(title="Select Browser Profile Folder (contains 'History' file)")
    
    if not folder_selected:
        print("No folder selected. Exiting.")
        return None
    
    # Check for History file in the selected folder
    history_path = os.path.join(folder_selected, "History")
    
    if not os.path.exists(history_path):
        # Try lowercase 'history' as well
        history_path = os.path.join(folder_selected, "history")
        if not os.path.exists(history_path):
            print(f"Error: No 'History' file found in {folder_selected}")
            return None
    
    return folder_selected  # Return folder path for extensions lookup


def main():
    # ANSI color codes
    CYAN = "\033[96m"
    GREEN = "\033[92m"
    YELLOW = "\033[93m"
    RED = "\033[91m"
    BOLD = "\033[1m"
    RESET = "\033[0m"
    
    try:
        # Get history file from user
        profile_folder = select_history_file()
        
        if not profile_folder:
            return
        
        # Build history path
        history_path = os.path.join(profile_folder, "History")
        if not os.path.exists(history_path):
            history_path = os.path.join(profile_folder, "history")
        
        print(f"{CYAN}▸ Using history file:{RESET} {history_path}")
        
        # Copy to temp location (browser locks the original)
        temp_db = copy_history_db(history_path)
        print(f"{CYAN}▸ Copied history database...{RESET}")
        
        # Parse history
        print(f"{CYAN}▸ Parsing history...{RESET}")
        data = parse_history(temp_db)
        
        # Parse extensions
        print(f"{CYAN}▸ Parsing extensions...{RESET}")
        data["extensions"] = parse_extensions(profile_folder)
        
        print(f"{GREEN}  ├─ History entries:   {len(data['history'])}{RESET}")
        print(f"{GREEN}  ├─ Search queries:    {len(data['searches'])}{RESET}")
        print(f"{GREEN}  ├─ Downloads:         {len(data['downloads'])}{RESET}")
        print(f"{GREEN}  └─ Extensions:       {len(data['extensions'])}{RESET}")
        
        # Save to same location as the script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        base_name = "browser_history.xlsx"
        output_path = os.path.join(script_dir, base_name)
        
        # Handle existing files by appending a number
        counter = 1
        while os.path.exists(output_path):
            name, ext = os.path.splitext(base_name)
            output_path = os.path.join(script_dir, f"{name}_{counter}{ext}")
            counter += 1
        
        # Export to XLSX
        export_to_xlsx(data, output_path)
        
        # Clean up temp file
        os.remove(temp_db)
        
        # Print clickable file path
        file_url = f"file://{output_path}"
        print()
        print(f"{GREEN}✓ Success!{RESET} Output saved to:")
        print(f"  {BOLD}{file_url}{RESET}")
        print()
        
    except FileNotFoundError as e:
        print(f"{RED}✗ Error:{RESET} {e}")
        print(f"\n{YELLOW}▸ Make sure you've selected the correct folder containing the History file.{RESET}")
    except Exception as e:
        print(f"{RED}✗ Error:{RESET} {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()